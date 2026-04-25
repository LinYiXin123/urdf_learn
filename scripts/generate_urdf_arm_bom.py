from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT_PATH = Path("output/spreadsheet/SO-ARM100_具身智能_BOM.xlsx")
FX_USD_TO_CNY = 6.898  # Approx. USD/CNY reference gathered in March 2026 web results.


@dataclass(frozen=True)
class MeshDimension:
    name: str
    min_mm: tuple[float, float, float]
    max_mm: tuple[float, float, float]
    extent_mm: tuple[float, float, float]


ARCHIVE_MESHES = [
    MeshDimension("Y1.stl", (-35.913006, -59.64102, -46.943775), (43.708035, 59.64102, 19.073833), (79.62104, 119.28204, 66.01761)),
    MeshDimension("Y2.stl", (-38.919006, -20.901382, -25.650286), (4.77384, 20.901382, 45.038994), (43.69285, 41.802765, 70.68928)),
    MeshDimension("Y4G.stl", (-29.45516, -23.870121, 166.46474), (4.201023, 24.744173, 277.12302), (33.65618, 48.614296, 110.65828)),
    MeshDimension("Y5.stl", (-29.493807, -23.780022, 283.35687), (4.1661935, 37.544395, 362.46417), (33.66, 61.324417, 79.1073)),
    MeshDimension("Y6G.stl", (-35.27938, -21.811815, 359.71524), (9.371094, 37.632515, 467.14572), (44.650475, 59.44433, 107.43048)),
    MeshDimension("Y7G.stl", (-36.433216, -23.02268, 370.32208), (10.557326, 5.5093265, 467.14572), (46.990543, 28.532007, 96.82364)),
    MeshDimension("YG.stl", (38.678135, -20.09372, -33.323204), (43.708035, 20.09372, 16.208055), (5.0298996, 40.18744, 49.531258)),
]

ARCHIVE_ASSEMBLY_MIN_MM = (-38.919006, -59.64102, -46.943775)
ARCHIVE_ASSEMBLY_MAX_MM = (43.708035, 59.64102, 467.14572)
ARCHIVE_ASSEMBLY_EXTENT_MM = (82.627045, 119.28204, 514.0895)
URDF_ASSEMBLY_EXTENT_CM = (8.286546, 11.928204, 51.500623)


def usd_to_cny(usd: float) -> float:
    return round(usd * FX_USD_TO_CNY, 2)


def mm_to_cm(values: Iterable[float]) -> list[float]:
    return [round(v / 10.0, 3) for v in values]


def style_header(ws, row: int, fill: str = "D9EAF7") -> None:
    thin = Side(style="thin", color="D0D7DE")
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.font = Font(bold=True, color="1F2328")
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)


def apply_common_sheet_style(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def add_hyperlink(cell, url: str, label: str = "购买链接") -> None:
    cell.value = label
    cell.hyperlink = url
    cell.style = "Hyperlink"


def build_dimensions_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "尺寸分析"
    ws.append(
        [
            "模型文件",
            "Min X (cm)",
            "Min Y (cm)",
            "Min Z (cm)",
            "Max X (cm)",
            "Max Y (cm)",
            "Max Z (cm)",
            "尺寸 X (cm)",
            "尺寸 Y (cm)",
            "尺寸 Z (cm)",
            "判断",
        ]
    )
    style_header(ws, 1)

    for mesh in ARCHIVE_MESHES:
        ws.append(
            [
                mesh.name,
                *mm_to_cm(mesh.min_mm),
                *mm_to_cm(mesh.max_mm),
                *mm_to_cm(mesh.extent_mm),
                "归档装配件，单位更像 mm，坐标已落在整机装配坐标系",
            ]
        )

    ws.append([])
    ws.append(
        [
            "归档整机包围盒",
            *mm_to_cm(ARCHIVE_ASSEMBLY_MIN_MM),
            *mm_to_cm(ARCHIVE_ASSEMBLY_MAX_MM),
            *mm_to_cm(ARCHIVE_ASSEMBLY_EXTENT_MM),
            "这套 STL 合并后的整机约 8.263 x 11.928 x 51.409 cm",
        ]
    )
    ws.append(
        [
            "当前 URDF 整机包围盒",
            "",
            "",
            "",
            "",
            "",
            "",
            *[round(v, 3) for v in URDF_ASSEMBLY_EXTENT_CM],
            "与归档 STL 尺寸高度一致，说明两者是同一量级的机械臂",
        ]
    )
    ws.append([])
    ws.append(["结论"])
    ws.append(
        [
            "这些 Y1/Y2/Y4G/Y5/Y6G/Y7G/YG 都是模型文件，而且是同一套机械臂的分件模型。"
        ]
    )
    ws.append(
        [
            "结合官方 SO-ARM100 BOM、打印说明和 ST3215 规格，这套模型最稳妥的落地方案就是按 SO-ARM100 单臂/双臂体系来配硬件。"
        ]
    )

    ws["A12"].font = Font(bold=True)
    ws["A13"].alignment = Alignment(wrap_text=True)
    ws["A14"].alignment = Alignment(wrap_text=True)
    apply_common_sheet_style(
        ws,
        {
            "A": 20,
            "B": 11,
            "C": 11,
            "D": 11,
            "E": 11,
            "F": 11,
            "G": 11,
            "H": 11,
            "I": 11,
            "J": 11,
            "K": 44,
        },
    )


def build_core_bom_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("核心BOM_单臂")
    headers = [
        "类别",
        "部件",
        "数量",
        "单价(RMB)",
        "小计(RMB)",
        "是否必须",
        "用途",
        "适配说明",
        "链接",
        "来源",
    ]
    ws.append(headers)
    style_header(ws, 1)

    rows = [
        [
            "执行器",
            "STS3215 串行总线舵机",
            6,
            97.72,
            None,
            "必须",
            "6 个关节执行器",
            "官方 SO-ARM100 单臂 BOM 指定 6 个；ST3215 外形约 45.2 x 24.7 x 35 mm，与这套打印件体系匹配",
            "https://item.taobao.com/item.htm?id=712179366565&skuId=5268252241438",
            "TheRobotStudio SO100.md",
        ],
        [
            "配件",
            "3-pin 舵机线",
            6,
            0.00,
            None,
            "必须",
            "总线串接和控制板到舵机的连接",
            "多数 ST3215 商品默认随舵机附带 15cm 线；若卖家未附带，需补 6 根",
            "https://item.taobao.com/item.htm?id=712179366565&skuId=5268252241438",
            "DFRobot/Waveshare 商品说明 + LeRobot 装配步骤",
        ],
        [
            "配件",
            "舵盘 + 舵盘固定螺丝",
            6,
            0.00,
            None,
            "必须",
            "把打印件固定到舵机输出轴",
            "多数 ST3215 商品默认随舵机附带；装配教程明确需要双舵盘和固定螺丝",
            "https://item.taobao.com/item.htm?id=712179366565&skuId=5268252241438",
            "LeRobot 装配步骤",
        ],
        [
            "控制",
            "总线舵机控制板",
            1,
            27.00,
            None,
            "必须",
            "电脑通过 USB 控制整条舵机总线",
            "官方 BOM 指定；Waveshare 控制板在 LeRobot 文档中被直接提及",
            "https://detail.tmall.com/item.htm?id=738817173460&skuId=5096283384143",
            "TheRobotStudio SO100.md",
        ],
        [
            "连接",
            "USB-C 数据线",
            1,
            23.90,
            None,
            "必须",
            "电脑连接控制板",
            "官方单臂 BOM 指定 1 根",
            "https://detail.tmall.com/item.htm?id=44425281296&skuId=5611379016222",
            "TheRobotStudio SO100.md",
        ],
        [
            "供电",
            "电源适配器",
            1,
            22.31,
            None,
            "必须",
            "给单臂供电",
            "SO-ARM100 Arm Kit 版本官方教程明确写 5V 电源；先按官方安全方案走",
            "https://item.taobao.com/item.htm?id=544824248494&skuId=4974994129990",
            "TheRobotStudio SO100.md + Seeed wiki",
        ],
        [
            "固定",
            "桌夹 2 只装",
            1,
            7.80,
            None,
            "必须",
            "把机械臂固定到桌面",
            "官方单臂 BOM 指定 2 个夹具",
            "https://detail.tmall.com/item.htm?id=738636473238&skuId=5505939904942",
            "TheRobotStudio SO100.md",
        ],
        [
            "结构",
            "3D 打印外壳/骨架",
            1,
            usd_to_cny(31.90),
            None,
            "必须",
            "机械臂外壳与连杆结构",
            "如果你不自己打印，Seeed 现成 SO-ARM100 打印件最省心；如果自己打印，可直接用项目里的 STL",
            "https://www.seeedstudio.com/SO-ARM100-3D-printed-Enclosure-p-6409.html",
            "Seeed 商品页",
        ],
        [
            "工具",
            "小号十字螺丝刀套装",
            1,
            14.90,
            None,
            "必须",
            "装配和清理支撑",
            "官方 BOM 把它列为单臂必须工具；新手建议直接备一套",
            "https://detail.tmall.com/item.htm?id=675684600845&skuId=4856851392176",
            "TheRobotStudio SO100.md",
        ],
    ]

    start_row = 2
    for row in rows:
        ws.append(row)
        current_row = ws.max_row
        ws[f"E{current_row}"] = f"=C{current_row}*D{current_row}"
        add_hyperlink(ws[f"I{current_row}"], row[8])

    summary_start = ws.max_row + 2
    ws[f"A{summary_start}"] = "官方单臂 BOM 小计（不含现成打印件）"
    ws[f"D{summary_start}"] = "RMB"
    ws[f"E{summary_start}"] = 682.23
    ws[f"A{summary_start + 1}"] = "按现成打印件直接落地的小计"
    ws[f"D{summary_start + 1}"] = "RMB"
    ws[f"E{summary_start + 1}"] = f"=SUM(E2:E{ws.max_row})"
    ws[f"A{summary_start + 2}"] = "推荐结论"
    ws[f"B{summary_start + 2}"] = "对新手最稳、最省事的方案：按这一页买齐，直接做一台单臂 follower。"
    ws[f"A{summary_start + 3}"] = "提示"
    ws[f"B{summary_start + 3}"] = "如果你准备自己打印，这一页里“3D 打印外壳/骨架”可以替换成自己的打印成本。"

    for row in (summary_start, summary_start + 1, summary_start + 2, summary_start + 3):
        ws[f"A{row}"].font = Font(bold=True)

    apply_common_sheet_style(
        ws,
        {
            "A": 10,
            "B": 22,
            "C": 8,
            "D": 12,
            "E": 12,
            "F": 10,
            "G": 18,
            "H": 42,
            "I": 10,
            "J": 26,
        },
    )


def build_embodied_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("具身扩展_双臂采集")
    headers = [
        "方案",
        "部件",
        "数量",
        "单价(RMB)",
        "小计(RMB)",
        "是否必须",
        "为什么需要",
        "链接",
        "来源",
    ]
    ws.append(headers)
    style_header(ws, 1, fill="E8F5E9")

    camera_price = usd_to_cny(16.66)
    jetson_price = usd_to_cny(249.00)

    rows = [
        [
            "必加",
            "第二套单臂核心硬件（leader 臂）",
            1,
            887.39,
            None,
            "必须",
            "如果你要做 LeRobot 式示教采集，单 follower 不够，还需要一台 leader 臂做 teleop",
            "",
            "由单臂核心 BOM 扣除可共享螺丝刀后估算",
        ],
        [
            "必加",
            "UVC/MJPG 1080p USB 相机",
            2,
            camera_price,
            None,
            "建议",
            "前视 + 侧视最实用；LeRobot/Seeed 文档给了双相机 OpenCV 配置示例",
            "https://geniuspycam.com/products/720p-1080p-4mp-5mp-8m-hd-usb-camera-module-mjpeg-30fps-high-speed-mini-cctv-linux-uvc-android-webcam-mini-surveillance-cam-audio",
            "Geniuspycam 商品页 + Seeed LeRobot 文档",
        ],
        [
            "必加",
            "现有电脑/笔记本",
            1,
            0.00,
            None,
            "必须",
            "最具性价比的做法是复用你现有电脑来配电机、采集数据和跑基础控制",
            "",
            "成本最优假设",
        ],
        [
            "可选",
            "Jetson Orin Nano Super Developer Kit",
            1,
            jetson_price,
            None,
            "可选",
            "如果你后面想做边缘端独立推理，而不是一直依赖电脑，可以再加这一块",
            "https://www.nvidia.com/en-us/autonomous-machines/embedded-systems/jetson-orin/nano-super-developer-kit/",
            "NVIDIA 官方页面",
        ],
    ]

    for row in rows:
        ws.append(row)
        current_row = ws.max_row
        ws[f"E{current_row}"] = f"=C{current_row}*D{current_row}"
        if row[7]:
            add_hyperlink(ws[f"H{current_row}"], row[7])

    summary_start = ws.max_row + 2
    ws[f"A{summary_start}"] = "双臂采集最小落地总价（复用现有电脑，不买 Jetson）"
    ws[f"E{summary_start}"] = f"=E2+E3+E4"
    ws[f"A{summary_start + 1}"] = "双臂采集 + Jetson 独立部署总价"
    ws[f"E{summary_start + 1}"] = f"=E2+E3+E4+E5"
    ws[f"A{summary_start + 2}"] = "建议"
    ws[f"B{summary_start + 2}"] = "新手先做单臂 follower -> 再补一台 leader -> 再补相机；Jetson 放到第三阶段最划算。"

    for row in (summary_start, summary_start + 1, summary_start + 2):
        ws[f"A{row}"].font = Font(bold=True)

    apply_common_sheet_style(
        ws,
        {
            "A": 10,
            "B": 28,
            "C": 8,
            "D": 12,
            "E": 12,
            "F": 10,
            "G": 38,
            "H": 10,
            "I": 24,
        },
    )


def build_sources_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("来源与假设")
    ws.append(["类型", "结论/假设", "来源 URL", "备注"])
    style_header(ws, 1, fill="FFF3CD")

    rows = [
        [
            "尺寸",
            "归档 STL 合并后的整机约 8.263 x 11.928 x 51.409 cm",
            "本地计算（归档/Y*.stl）",
            "这些 STL 的坐标看起来已在装配坐标系中，单位更像 mm",
        ],
        [
            "匹配性",
            "这套模型与 SO-ARM100 的尺寸级别和硬件体系高度吻合",
            "https://raw.githubusercontent.com/TheRobotStudio/SO-ARM100/main/SO100.md",
            "核心 BOM 和打印说明都指向同一类结构",
        ],
        [
            "打印适配",
            "官方提供 STS3215 量规，说明打印件就是围绕 STS3215 设计",
            "https://raw.githubusercontent.com/TheRobotStudio/SO-ARM100/main/SO100.md",
            "Gauge Zero / Gauge Tight 用来校验 STS3215 配合",
        ],
        [
            "舵机尺寸",
            "ST3215 外形约 45.2 x 24.7 x 35 mm",
            "https://www.waveshare.com/st3215-servo.htm",
            "搜索结果和商品说明一致",
        ],
        [
            "舵机附件",
            "ST3215 商品通常含 1 根 3-pin 线、舵盘和螺丝",
            "https://www.dfrobot.com/product-2961.html",
            "不同卖家附件可能有差异，因此 BOM 里单独列了这两项并标 0 元核对",
        ],
        [
            "单臂官方成本",
            "官方 SO-ARM100 单臂人民币 BOM 合计为 682.23",
            "https://raw.githubusercontent.com/TheRobotStudio/SO-ARM100/main/SO100.md",
            "这个数字不含你自己本地打印或买现成打印件的差异",
        ],
        [
            "打印件直购",
            "Seeed 现成 SO-ARM100 打印件价格为 31.90 美元",
            "https://www.seeedstudio.com/SO-ARM100-3D-printed-Enclosure-p-6409.html",
            "已按约 1 USD = 6.898 CNY 近似换算",
        ],
        [
            "汇率",
            "USD/CNY 采用近似 6.898",
            "https://www.currency-converter.org.uk/currency-rates/historical/table/USD-CNY.html",
            "仅用于把美元商品换算成人民币参考值",
        ],
        [
            "具身采集",
            "LeRobot 风格的数据采集需要 leader + follower，且常配 1 到 2 个以上相机",
            "https://huggingface.co/docs/lerobot/en/so100",
            "Seeed/LeRobot 文档里给出了双臂和双相机命令示例",
        ],
        [
            "训练算力",
            "本地训练示例使用 CUDA 设备",
            "https://wiki.seeedstudio.com/lerobot_so100m_new/",
            "因此真正本地训练通常需要 NVIDIA GPU 或 Jetson；入门阶段建议先复用现有电脑做采集",
        ],
        [
            "Jetson 可选项",
            "Jetson Orin Nano Super 官方价 249 美元",
            "https://www.nvidia.com/en-us/autonomous-machines/embedded-systems/jetson-orin/nano-super-developer-kit/",
            "属于第三阶段升级项，不是装臂必需品",
        ],
    ]

    for row in rows:
        ws.append(row)
        if row[2].startswith("http"):
            add_hyperlink(ws[f"C{ws.max_row}"], row[2], label=row[2])

    ws.append([])
    ws.append(["诚实说明"])
    ws.append(
        [
            "我不能对“100% 一定能装起来”做绝对保证，因为这需要剖开 CAD 腔体、核对每个卖家的附件、并实际试装。"
        ]
    )
    ws.append(
        [
            "但基于 STL 实测尺寸、SO-ARM100 官方 BOM、官方 ST3215 量规、LeRobot/Seeed 装配文档，这已经是我能给出的最稳妥、最接近可直接落地的一套高性价比方案。"
        ]
    )
    ws["A14"].font = Font(bold=True)
    ws["A15"].alignment = Alignment(wrap_text=True)
    ws["A16"].alignment = Alignment(wrap_text=True)

    apply_common_sheet_style(
        ws,
        {
            "A": 12,
            "B": 36,
            "C": 58,
            "D": 38,
        },
    )


def build_review_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("装配审查")
    ws.append(["检查项", "结论", "当前状态", "需要你确认/补充", "依据"])
    style_header(ws, 1, fill="FDE2E4")

    rows = [
        [
            "这份 Excel 是否已是最终可下单版",
            "不是最终版采购清单，更准确地说它是“官方兼容核心 BOM + 具身扩展路线图”。",
            "部分完整",
            "还需要核对你买到的 ST3215 是否带齐配件包，以及你是否自己打印。",
            "SO100 官方 BOM 只列核心件，不单列紧固件。",
        ],
        [
            "单臂能否实现具身智能",
            "单臂能装成可控机械臂，但不能独立完成标准 LeRobot leader/follower 示教采集闭环。",
            "不能单独完成标准采集",
            "若要标准具身数据采集，至少再补 1 台 leader 臂、相机和主机算力。",
            "LeRobot SO100/SO101 文档默认是 leader + follower 架构。",
        ],
        [
            "核心主臂是否缺少单独轴承",
            "按官方 SO100/SO101 核心 BOM，不需要额外单独买散装轴承。",
            "看起来不缺",
            "注意 ST3215 伺服本体内部自带 ball bearing，但不是外置散件。",
            "官方 BOM 未列散装轴承；ST3215 规格写有 ball bearing。",
        ],
        [
            "核心主臂是否缺少单独螺母",
            "核心主臂官方 BOM 未要求额外散装螺母；只有某些可选相机支架才额外需要 M3 螺母。",
            "主臂本体大概率不缺",
            "如果你准备加 SO101 Hex-Nut 相机支架，额外补 M3 螺母即可。",
            "SO101 wrist hex-nut camera mount README。",
        ],
        [
            "核心主臂是否缺少螺丝尺寸信息",
            "缺。当前 Excel 没把“主臂依赖舵机附件包里的螺丝”写清楚。",
            "需要补说明",
            "最少要核对每个舵机都带 servo horn screw accessory kit；建议另外备一套 M3x6 和 M2.5x4。",
            "DFRobot shipping list + SO101 STEP 装配体。",
        ],
        [
            "推荐的备用紧固件",
            "M3x6 盘头螺丝、M2.5x4 小盘头螺丝、M3 六角螺母。",
            "建议额外准备",
            "推荐备货量：M3x6 至少 20-50 颗，M2.5x4 至少 8-20 颗，M3 螺母 4-10 个。",
            "SO101 Assembly.step 中可见 M3x6、M2.5x4、M3 nut；相机可选件也复用这些规格。",
        ],
        [
            "舵机自带附件是否关键",
            "是。官方可选件 README 多次写到‘these came with your Feetech servos’。",
            "高风险依赖",
            "如果卖家卖的是裸机，不带线材/舵盘/螺丝包，这份 BOM 就不完整。",
            "Wrist cam / overhead cam / D435 mount README。",
        ],
        [
            "打印件与 ST3215 尺寸是否同量级",
            "是。你的归档整机约 8.263 x 11.928 x 51.409 cm，和本地 URDF 约 8.287 x 11.928 x 51.501 cm 高度一致。",
            "通过尺度检查",
            "这只能证明尺度匹配，不能替代实际卡尺测量和试装。",
            "本地 STL/URDF 实测。",
        ],
        [
            "ST3215 能否物理嵌入这些件",
            "从尺寸等级看高度可行，但我不能仅凭外包围盒就给出 100% 嵌入保证。",
            "高概率可行，未实装",
            "务必先打印 Gauge_0 / Gauge_tight_1，并只买 1-2 个舵机做试装。",
            "官方提供 STS3215 gauge，专门用于验证打印配合。",
        ],
        [
            "当前最稳的落地顺序",
            "先打印量规 -> 先买 1-2 个 ST3215 试装 -> 再补齐 6 个 follower -> 最后再做 leader 和相机。",
            "推荐流程",
            "不要第一次就全量下单双臂和相机。",
            "基于装配风险和具身扩展路径的综合建议。",
        ],
    ]

    for row in rows:
        ws.append(row)

    ws.append([])
    ws.append(["我对这份方案的最终判断"])
    ws.append(
        [
            "对“先做一台能装起来的机械臂”来说，这份 Excel 的主路线是对的；对“直接一次买齐完整具身系统”来说，它还缺装配级细项和卖家附件核对。"
        ]
    )

    ws["A13"].font = Font(bold=True)
    ws["A14"].alignment = Alignment(wrap_text=True)

    apply_common_sheet_style(
        ws,
        {
            "A": 20,
            "B": 36,
            "C": 16,
            "D": 36,
            "E": 28,
        },
    )


def build_domestic_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("国内采购推荐")
    ws.append(["方案", "推荐度", "适合你吗", "核心配置", "国内购买入口", "价格/可得性", "为什么推荐", "不推荐什么"])
    style_header(ws, 1, fill="DFF6DD")

    rows = [
        [
            "A. 最省钱可落地：SO-ARM100 单臂 DIY",
            "最高",
            "适合已有电脑、能自己打印，先做第一台机械臂的人",
            "STS3215/ ST3215 7.4V C001 x6 + 舵机驱动板 x1 + 5V 电源 x1 + USB-C x1 + 桌夹 x1",
            "舵机: https://item.taobao.com/item.htm?id=712179366565&skuId=5268252241438\n驱动板: https://detail.tmall.com/item.htm?id=738817173460&skuId=5096283384143\nUSB-C: https://detail.tmall.com/item.htm?id=44425281296&skuId=5611379016222\n电源: https://item.taobao.com/item.htm?id=544824248494&skuId=4974994129990\n桌夹: https://detail.tmall.com/item.htm?id=738636473238&skuId=5505939904942",
            "官方 CN BOM 可验证，单臂合计约 682.23 RMB，不含你自己打印成本",
            "这是国内最便宜、最稳、和你当前 STL/URDF 最匹配的一条路。官方明确把这套作为 CN BOM 给出，而且你这套模型尺寸和 SO-ARM100 同量级。",
            "不推荐一开始买 Jetson 套餐，也不推荐改众灵/普通 PWM。",
        ],
        [
            "B. 标准具身采集最省钱：SO-ARM100 双臂 DIY",
            "很高",
            "适合要做 LeRobot 标准 leader/follower 采集的人",
            "在 A 的基础上翻倍成 2 台臂，再加 1 到 2 个 UVC 相机和现有电脑",
            "官方 CN BOM 见: https://raw.githubusercontent.com/TheRobotStudio/SO-ARM100/main/SO100.md\nSeeed 中文教程: https://wiki.seeedstudio.com/cn/lerobot_so100m/",
            "官方 CN BOM 两臂约 1343.16 RMB，不含打印件；相机建议单独补",
            "这是目前国内最具性价比的“标准具身”起点。单臂只能做可控机械臂，双臂才是官方示教采集闭环。",
            "不推荐先买整套 Jetson，先用自己电脑更省钱。",
        ],
        [
            "C. 无打印机最省事：国内正规套件",
            "高",
            "适合没有打印条件，但想快速落地的人",
            "优先 SO-ARM100 3D 打印套餐；如果还想一步到位采集，选双视角摄像头套餐",
            "微雪国内站: https://www.waveshare.net/shop/SO-ARM100-3DP-Parts-Kit.htm\nSeeed 中国淘宝入口(官方仓库给出): https://item.taobao.com/item.htm?id=878010637397&skuId=5915703371829&spm=a213gs.v2success.0.0.4cbf4831mkqWLn",
            "价格需进店查看实时价；货源稳定、兼容性高",
            "微雪页面明确把 SO-ARM100 系列分成“无 3D 打印套餐 / 3D 打印套餐 / 双视角摄像头套餐 / Jetson 套件套餐”。如果你不想自己找打印件和相机支架，这条路很省心。",
            "不推荐直接买 Jetson 套件，除非你已经确定要本地边缘推理。",
        ],
        [
            "D. 完全组装/企业店方案",
            "中高",
            "适合预算更宽、想少折腾装配的人",
            "整机或已装好套件",
            "WowRobo 中国: https://item.taobao.com/item.htm?ft=t&id=860171734711\nNeoBot 中国: https://item.taobao.com/item.htm?ft=t&id=957685951340",
            "价格多半高于 DIY；实时价需进店确认",
            "官方仓库把它们列为中国区套件来源，说明兼容性路径是可信的。",
            "不推荐把它当最低预算方案，它更适合省时间，不适合省钱。",
        ],
        [
            "电机型号结论",
            "必须看",
            "适合你当前这套模型",
            "SO-ARM100 / 你当前 STL 路线优先 7.4V STS3215/ ST3215 C001（约 19.5kg.cm）",
            "官方说明: https://wiki.seeedstudio.com/cn/lerobot_so100m/\n微雪说明: https://www.waveshare.net/shop/SO-ARM100-3DP-Parts-Kit.htm",
            "标准版就够入门，专业版/12V 更贵",
            "官方明确写了 SO-ARM100 使用 7.4V 堵转扭矩 19.5kg.cm 的总线舵机；同时说明常规教学/项目练手选 SO-ARM100，更大抓取力再上 SO-ARM101 30kg.cm。",
            "不推荐为了‘更大 kg’直接切 SO-ARM101 或 12V 舵机，因为你的现有模型与 SO-ARM100 更吻合，且总成本和供电复杂度会抬高。",
        ],
    ]

    for row in rows:
        ws.append(row)
        current_row = ws.max_row
        links = row[4].split("\n")
        if links:
            first_url = links[0].split(": ", 1)[-1] if ": " in links[0] else links[0]
            if first_url.startswith("http"):
                add_hyperlink(ws[f"F{current_row}"], first_url, label="入口链接")

    ws.append([])
    ws.append(["我的最终筛选"])
    ws.append(
        [
            "如果你追求‘最划算且能落地’，选 A；如果你追求‘真做具身数据采集’，在 A 跑通后升级到 B；如果你没有打印条件，选 C；如果你只想少折腾装配，选 D。"
        ]
    )
    ws["A8"].font = Font(bold=True)
    ws["A9"].alignment = Alignment(wrap_text=True)

    apply_common_sheet_style(
        ws,
        {
            "A": 24,
            "B": 10,
            "C": 24,
            "D": 28,
            "E": 54,
            "F": 10,
            "G": 32,
            "H": 28,
        },
    )


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    build_dimensions_sheet(wb)
    build_core_bom_sheet(wb)
    build_embodied_sheet(wb)
    build_sources_sheet(wb)
    build_review_sheet(wb)
    build_domestic_sheet(wb)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.column <= 3:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                elif cell.data_type != "f":
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            if col_letter in ("D", "E"):
                for cell in ws[col_letter]:
                    if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                        cell.number_format = '0.00'

    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH.resolve())


if __name__ == "__main__":
    main()
