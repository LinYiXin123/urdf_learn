from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


OUTPUT_PATH = Path("output/spreadsheet/SO-ARM100_最推荐_无打印机方案.xlsx")


def style_header(ws, row: int, fill: str = "D9EAF7") -> None:
    thin = Side(style="thin", color="D0D7DE")
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.font = Font(bold=True, color="1F2328")
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)


def apply_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def add_hyperlink(cell, url: str, label: str = "链接") -> None:
    cell.value = label
    cell.hyperlink = url
    cell.style = "Hyperlink"


def build_summary_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "推荐方案"
    ws.append(["项目", "数量", "单价(RMB)", "小计(RMB)", "是否计入总价", "说明", "购买入口"])
    style_header(ws, 1, fill="DFF6DD")

    rows = [
        [
            "SO-ARM100 CAM Kit",
            1,
            5523.69,
            None,
            "计入",
            "我最推荐的无打印机方案。官方说明它适合“需要多视角数据采集且已有 Jetson Orin Nano 主控板或自有电脑的用户”，并且页面明确写双视角摄像头套餐为推荐项。",
            "https://www.waveshare.net/shop/SO-ARM100-CAM-Kit.htm",
        ],
        [
            "现有电脑/笔记本",
            1,
            0.00,
            None,
            "计入",
            "按你已有电脑来算。微雪页面明确写双视角摄像头套餐可搭配 PC 使用，因此不强制你买 Jetson。",
            "https://www.waveshare.net/shop/SO-ARM100-CAM-Kit.htm",
        ],
        [
            "打印机",
            0,
            0.00,
            0.00,
            "已覆盖",
            "你没有打印机，这个方案已经把树脂 3D 打印结构件算进套件里了，不需要另买打印机或单独找打印服务。",
            "",
        ],
        [
            "额外 Jetson Orin Nano 套件",
            0,
            0.00,
            0.00,
            "不计入",
            "不建议第一批购买。你要先把双臂、双相机、数据采集跑通，再决定是否上边缘推理主机。",
            "https://www.waveshare.net/shop/SO-ARM100-CAM-Kit.htm",
        ],
    ]

    for row in rows:
        ws.append(row)
        r = ws.max_row
        if row[3] is None:
            ws[f"D{r}"] = f"=B{r}*C{r}"
        if row[6]:
            add_hyperlink(ws[f"G{r}"], row[6], "购买页")

    ws["A7"] = "总价（按已有电脑计算）"
    ws["D7"] = "=SUMIF(E2:E5,\"计入\",D2:D5)"
    ws["A8"] = "一句话结论"
    ws["B8"] = "如果你没有打印机，又是按我最推荐的路线来做具身智能机械臂，直接买 SO-ARM100 CAM Kit，然后用你现有电脑开始做 leader/follower + 双视角采集。"
    ws["A9"] = "为什么不是 3DP Parts Kit"
    ws["B9"] = "3DP Parts Kit 更便宜，但还得自己补相机和采集配件；而你明确是要做具身智能，这一步我更推荐直接走 CAM Kit。"
    ws["A10"] = "为什么不是 Jetson Kit"
    ws["B10"] = "Jetson 套件不是第一批最划算选择。先用 PC 跑通数据采集和回放，后面再升级算力更稳。"

    for cell in ("A7", "A8", "A9", "A10"):
        ws[cell].font = Font(bold=True)
    for row in range(2, 11):
        for cell in ws[row]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    apply_widths(
        ws,
        {
            "A": 24,
            "B": 8,
            "C": 12,
            "D": 12,
            "E": 12,
            "F": 48,
            "G": 10,
        },
    )


def build_detail_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("套件解读")
    ws.append(["检查项", "结论", "依据"])
    style_header(ws, 1, fill="FFF3CD")

    rows = [
        [
            "它是不是按具身智能准备好的",
            "是。页面明确说 CAM Kit 配备双视角摄像头、USB HUB，并且适用于需要多视角数据采集的用户。",
            "https://www.waveshare.net/shop/SO-ARM100-CAM-Kit.htm",
        ],
        [
            "它是不是按双臂算的",
            "是。SO-ARM100 系列配置写的是 ST3215-7.4V Servo × 12，这正对应 leader + follower 两台机械臂。",
            "https://www.waveshare.com/so-arm100-3dp-parts-kit.htm",
        ],
        [
            "它的舵机是否适合你当前模型路线",
            "是。SO-ARM100 系列明确是 7.4V、19.5kg.cm、12 位磁编码器的总线舵机体系，与你当前 SO-ARM100 尺寸路线一致。",
            "https://www.waveshare.net/shop/SO-ARM100-CAM-Kit.htm",
        ],
        [
            "它有没有把打印件算进去",
            "有。CAM Kit 属于 SO-ARM100 系列里包含 3D 光敏树脂打印套件的版本。",
            "https://www.waveshare.com/so-arm100-3dp-parts-kit.htm",
        ],
        [
            "还缺什么",
            "套件页写的是可搭配 PC 或 Jetson Orin Nano 主控板使用，所以至少还需要一台可用电脑。按本表默认，你已有电脑，成本按 0 计。",
            "https://www.waveshare.net/shop/SO-ARM100-CAM-Kit.htm",
        ],
    ]

    for row in rows:
        ws.append(row)
        add_hyperlink(ws[f"C{ws.max_row}"], row[2], row[2])

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    apply_widths(ws, {"A": 24, "B": 52, "C": 56})


def build_compare_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("备选对比")
    ws.append(["方案", "国内参考价", "适合谁", "我是否推荐", "原因"])
    style_header(ws, 1, fill="FDE2E4")

    rows = [
        [
            "SO-ARM100 3DP Parts Kit",
            2149.28,
            "没有打印机，但只想先把双臂硬件装起来的人",
            "次选",
            "它把 3D 打印件算进去了，但不含双相机采集配件。你既然目标是具身智能，我更推荐直接上 CAM Kit。",
        ],
        [
            "SO-ARM100 CAM Kit",
            5523.69,
            "没有打印机，而且明确要做具身数据采集的人",
            "最推荐",
            "双臂 + 3D 打印件 + 双相机 + USB HUB，一步到位接近官方推荐采集环境，同时又不强制你买 Jetson。",
        ],
        [
            "SO-ARM100 Jetson Orin Nano Kit",
            None,
            "预算充足，且一开始就要本地边缘推理的人",
            "不作为首推",
            "页面存在价格映射疑似异常，我这次不拿它做推荐；而且你先用 PC 跑通更省钱、更稳。",
        ],
    ]

    for row in rows:
        ws.append(row)
        r = ws.max_row
        if row[1] is not None:
            ws[f"B{r}"] = row[1]

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    apply_widths(ws, {"A": 26, "B": 14, "C": 28, "D": 12, "E": 44})


def build_sources_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("来源")
    ws.append(["来源类型", "结论", "链接"])
    style_header(ws, 1, fill="E8F5E9")

    rows = [
        [
            "Waveshare 国内 CAM Kit 页面",
            "页面明确写了 SO-ARM100 CAM Kit、双视角摄像头套餐、可搭配 PC、适合多视角数据采集用户。",
            "https://www.waveshare.net/shop/SO-ARM100-CAM-Kit.htm",
        ],
        [
            "Waveshare 国内商品页",
            "SO-ARM100 / 101 系列介绍里明确写 SO-ARM100 为 19.5kg.cm 总线舵机、双视角套餐为推荐项。",
            "https://www.waveshare.net/shop/SO-ARM100-3DP-Parts-Kit.htm",
        ],
        [
            "Waveshare 国内分类页",
            "分类页可抓到 SO-ARM100 No-3DP Parts Kit 约 1698.82 RMB、3DP Parts Kit 约 2149.28 RMB、CAM Kit 约 5523.69 RMB。",
            "https://www.waveshare.net/left_column/Raspberry-Pi-Robotics.htm",
        ],
        [
            "官方 SO-100 BOM",
            "官方双臂 CN BOM 是 1343.16 RMB（不含打印件和相机套餐化打包差异），用于校验 Waveshare 套件路线的合理性。",
            "https://raw.githubusercontent.com/TheRobotStudio/SO-ARM100/main/SO100.md",
        ],
    ]

    for row in rows:
        ws.append(row)
        add_hyperlink(ws[f"C{ws.max_row}"], row[2], row[2])

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    apply_widths(ws, {"A": 20, "B": 54, "C": 60})


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    build_summary_sheet(wb)
    build_detail_sheet(wb)
    build_compare_sheet(wb)
    build_sources_sheet(wb)
    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH.resolve())


if __name__ == "__main__":
    main()
