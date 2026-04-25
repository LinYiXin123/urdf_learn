from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


OUTPUT_PATH = Path("output/spreadsheet/自有模型_第三方打印_具身落地方案.xlsx")


PRINT_SERVICE_PER_SET_RMB = 450.46
CAMERA_PER_UNIT_RMB = 98.00
SPARE_FASTENER_PACK_SINGLE_RMB = 30.00
SPARE_FASTENER_PACK_DUAL_RMB = 50.00


def style_header(ws, row: int, fill: str = "D9EAF7") -> None:
    thin = Side(style="thin", color="D0D7DE")
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.font = Font(bold=True, color="1F2328")
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)


def apply_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def add_hyperlink(cell, url: str, label: str = "链接") -> None:
    cell.value = label
    cell.hyperlink = url
    cell.style = "Hyperlink"


def build_summary_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "总览"
    ws.append(["方案", "总价(RMB)", "是否完整具身", "是否使用你的模型", "核心结论"])
    style_header(ws, 1, fill="DFF6DD")

    ws.append([
        "单臂落地版",
        1260.69,
        "否",
        "是",
        "用你的一套模型第三方打印，配 6 个 STS3215，就能落地一台可控视觉机械臂。",
    ])
    ws.append([
        "双臂具身采集版",
        2501.68,
        "是",
        "是",
        "用你的模型打印两套，加 12 个 STS3215、2 块控制板、2 个相机和现有电脑，就能搭出标准具身采集原型。",
    ])
    ws.append([
        "重要提醒",
        "",
        "",
        "",
        "你当前给我的模型更像 follower 结构。若要做标准 leader/follower，第二套模型可以先作为 leader 原型使用，但手感和人体工学大概率不如官方 leader 专用打印件。",
    ])

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    apply_widths(ws, {"A": 18, "B": 12, "C": 12, "D": 12, "E": 56})


def build_model_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("模型与打印")
    ws.append(["检查项", "结果", "含义/建议", "依据"])
    style_header(ws, 1, fill="FFF3CD")

    rows = [
        [
            "整机包围盒",
            "约 8.263 x 11.928 x 51.409 cm",
            "这说明你的模型是半米级桌面机械臂，不是缩放错误。",
            "本地对归档/Y*.stl 的几何计算",
        ],
        [
            "模型总体积",
            "约 302.70 cm^3（实体体积）",
            "用于估算第三方打印价格和材料用量。",
            "本地 STL 体积计算",
        ],
        [
            "FDM 实际材料估算",
            "约 113 g 到 131 g PLA+",
            "按 30% 到 35% 实体等效估算，符合官方 13% 到 15% infill 的轻量打印路线。",
            "本地计算 + 官方打印建议",
        ],
        [
            "推荐打印工艺",
            "FDM，PLA+，0.4 喷嘴，0.2 层高，15% infill",
            "如果第三方能稳定打印 Tough Resin 也可，但从成本和可修补性出发，先选 PLA+ 更稳。",
            "https://raw.githubusercontent.com/TheRobotStudio/SO-ARM100/main/SO100.md",
        ],
        [
            "第三方打印成本预算",
            "单套约 450.46 RMB，建议按 350 到 600 RMB 报价区间询价",
            "这个数值来自官方国内站 SO-ARM100 无 3D 打印套餐与 3D 打印套餐的价差，可作为你自有模型外包打印的现实参考。",
            "https://www.waveshare.net/left_column/Raspberry-Pi-Robotics.htm",
        ],
        [
            "打印前必须做的校验",
            "先打印 STS3215 Gauge_0 / Gauge_tight_1 或要求商家先打印一个舵机安装段试样",
            "这一步比盲目全量打印更重要，能显著降低‘装不进去’的风险。",
            "https://raw.githubusercontent.com/TheRobotStudio/SO-ARM100/main/SO100.md",
        ],
    ]

    for row in rows:
        ws.append(row)
        if row[3].startswith("http"):
            add_hyperlink(ws[f"D{ws.max_row}"], row[3], row[3])

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    apply_widths(ws, {"A": 20, "B": 22, "C": 48, "D": 56})


def build_hardware_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("硬件尺寸要求")
    ws.append(["硬件", "推荐规格", "关键尺寸/数量", "为什么这样选", "购买入口/来源"])
    style_header(ws, 1, fill="E8F5E9")

    rows = [
        [
            "关节舵机",
            "STS3215 / ST3215 7.4V C001",
            "45.2 x 24.7 x 35 mm；6 个/臂",
            "这是和你当前模型路线最接近、官方 CN BOM 可验证、并且支持总线回读与 12 位磁编码器的一类舵机。",
            "https://item.taobao.com/item.htm?id=712179366565&skuId=5268252241438",
        ],
        [
            "舵机附件包",
            "必须带 3-pin 线、舵盘、舵盘螺丝",
            "1 套/舵机",
            "你现在最容易卡死装配的不是舵机本体，而是卖家卖裸机不送附件包。",
            "https://www.dfrobot.com/product-2961.html",
        ],
        [
            "控制板",
            "官方推荐总线舵机控制板",
            "1 块/臂",
            "你现有前端和官方路线都更接近飞特总线舵机生态，优先走这条最稳。",
            "https://detail.tmall.com/item.htm?id=738817173460&skuId=5096283384143",
        ],
        [
            "电源",
            "5V 电源适配器（按官方 CN BOM）",
            "1 个/臂",
            "这是官方 CN BOM 已验证过的搭配，先按它落地，后续再看扭矩是否需要升级供电。",
            "https://item.taobao.com/item.htm?id=544824248494&skuId=4974994129990",
        ],
        [
            "USB 线",
            "USB-C 数据线",
            "1 根/臂",
            "控制板接电脑调试和运行。",
            "https://detail.tmall.com/item.htm?id=44425281296&skuId=5611379016222",
        ],
        [
            "固定件",
            "桌夹",
            "2 个/臂",
            "桌面机械臂如果底座不稳，视觉和控制都会一起变差。",
            "https://detail.tmall.com/item.htm?id=738636473238&skuId=5505939904942",
        ],
        [
            "相机",
            "UVC 摄像头，720p30 起步，优先 1080p30；推荐罗技 C270 或同级",
            "单臂 1 个，具身双臂 2 个",
            "用桌面 UVC webcam 最容易落地，不依赖额外腕部相机支架建模。",
            "https://www.jd.com/phb/key_6703325379caeabf8ee.html",
        ],
        [
            "备用紧固件",
            "M3x6 盘头、M2.5x4 小盘头、M3 六角螺母",
            "单臂备一包，双臂备两包",
            "你的模型能否装起来，和这些小东西是否齐全关系非常大。",
            "https://github.com/TheRobotStudio/SO-ARM100",
        ],
        [
            "采集主机",
            "现有 PC / 笔记本",
            "至少 4 个 USB 口，16GB 内存更稳",
            "做数据采集、标定、回放和基础控制不一定要 Jetson；先用 PC 最省钱。",
            "https://www.waveshare.net/shop/SO-ARM100-CAM-Kit.htm",
        ],
        [
            "训练主机",
            "可后续升级到 RTX 3060 12GB / RTX 4060 8GB 级别台式机，或用云 GPU",
            "非首批必须",
            "把机械和采集跑通优先于本地训练硬件采购。",
            "https://wiki.seeedstudio.com/cn/lerobot_so100m/",
        ],
    ]

    for row in rows:
        ws.append(row)
        if row[4].startswith("http"):
            add_hyperlink(ws[f"E{ws.max_row}"], row[4], "来源/入口")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    apply_widths(ws, {"A": 14, "B": 22, "C": 20, "D": 36, "E": 16})


def build_single_arm_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("单臂落地BOM")
    ws.append(["项目", "数量", "单价(RMB)", "小计(RMB)", "推荐理由", "链接"])
    style_header(ws, 1, fill="D9EAF7")

    rows = [
        ["第三方打印你的模型", 1, PRINT_SERVICE_PER_SET_RMB, None, "没有打印机时，用官方 3DP 套餐价差估算最现实。", ""],
        ["STS3215 / ST3215 7.4V C001", 6, 97.72, None, "与你当前模型路线最匹配的总线舵机。", "https://item.taobao.com/item.htm?id=712179366565&skuId=5268252241438"],
        ["总线舵机控制板", 1, 27.00, None, "官方 CN BOM。", "https://detail.tmall.com/item.htm?id=738817173460&skuId=5096283384143"],
        ["USB-C 数据线", 1, 23.90, None, "控制板连电脑。", "https://detail.tmall.com/item.htm?id=44425281296&skuId=5611379016222"],
        ["电源适配器", 1, 22.31, None, "先按官方 CN BOM 的已验证配置走。", "https://item.taobao.com/item.htm?id=544824248494&skuId=4974994129990"],
        ["桌夹 2 只装", 1, 7.80, None, "固定底座。", "https://detail.tmall.com/item.htm?id=738636473238&skuId=5505939904942"],
        ["小号十字螺丝刀套装", 1, 14.90, None, "新手建议备上。", "https://detail.tmall.com/item.htm?id=675684600845&skuId=4856851392176"],
        ["备用紧固件包", 1, SPARE_FASTENER_PACK_SINGLE_RMB, None, "防止舵机附件缺失或装配中丢件。", ""],
        ["UVC 摄像头", 1, CAMERA_PER_UNIT_RMB, None, "单臂做视觉观测和调试。", "https://www.jd.com/phb/key_6703325379caeabf8ee.html"],
        ["现有电脑", 1, 0.00, None, "按你已有电脑计算。", ""],
    ]

    for row in rows:
        ws.append(row)
        r = ws.max_row
        ws[f"D{r}"] = f"=B{r}*C{r}"
        if row[5]:
            add_hyperlink(ws[f"F{r}"], row[5], "入口")

    ws["A12"] = "总价"
    ws["D12"] = "=SUM(D2:D11)"
    ws["A13"] = "适用范围"
    ws["B13"] = "能落地一台可控视觉机械臂，但还不是标准 leader/follower 具身采集闭环。"

    ws["A12"].font = Font(bold=True)
    ws["A13"].font = Font(bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    apply_widths(ws, {"A": 22, "B": 8, "C": 12, "D": 12, "E": 34, "F": 10})


def build_dual_arm_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("完整具身BOM")
    ws.append(["项目", "数量", "单价(RMB)", "小计(RMB)", "说明", "链接"])
    style_header(ws, 1, fill="FDE2E4")

    rows = [
        ["第三方打印你的模型", 2, PRINT_SERVICE_PER_SET_RMB, None, "打印两套，分别做 follower 和 leader 原型。", ""],
        ["STS3215 / ST3215 7.4V C001", 12, 97.72, None, "双臂总计 12 个。", "https://item.taobao.com/item.htm?id=712179366565&skuId=5268252241438"],
        ["总线舵机控制板", 2, 27.00, None, "每臂 1 块。", "https://detail.tmall.com/item.htm?id=738817173460&skuId=5096283384143"],
        ["USB-C 数据线", 2, 23.90, None, "每臂 1 根。", "https://detail.tmall.com/item.htm?id=44425281296&skuId=5611379016222"],
        ["电源适配器", 2, 22.31, None, "每臂 1 个。", "https://item.taobao.com/item.htm?id=544824248494&skuId=4974994129990"],
        ["桌夹", 4, 5.20, None, "按双臂 4 个夹具估算。", "https://detail.tmall.com/item.htm?id=801399113134&skuId=5633627126649"],
        ["小号十字螺丝刀套装", 1, 14.90, None, "共享 1 套即可。", "https://detail.tmall.com/item.htm?id=675684600845&skuId=4856851392176"],
        ["备用紧固件包", 1, SPARE_FASTENER_PACK_DUAL_RMB, None, "双臂建议至少备一整包。", ""],
        ["UVC 摄像头", 2, CAMERA_PER_UNIT_RMB, None, "双视角采集。", "https://www.jd.com/phb/key_6703325379caeabf8ee.html"],
        ["现有电脑", 1, 0.00, None, "按你已有电脑计算。", ""],
    ]

    for row in rows:
        ws.append(row)
        r = ws.max_row
        ws[f"D{r}"] = f"=B{r}*C{r}"
        if row[5]:
            add_hyperlink(ws[f"F{r}"], row[5], "入口")

    ws["A12"] = "总价"
    ws["D12"] = "=SUM(D2:D11)"
    ws["A13"] = "是否能完成具身采集"
    ws["B13"] = "可以作为完整具身采集原型。前提是你愿意把第二套模型当 leader 原型使用，或后续再优化 leader 手柄/人机交互结构。"
    ws["A14"] = "我最诚实的提醒"
    ws["B14"] = "从硬件数量和成本上，这版是能落地的；但从人体工学上，你当前模型并不是为 leader 手感专门优化的，所以‘能用’不代表‘示教最好用’。"

    for cell in ("A12", "A13", "A14"):
        ws[cell].font = Font(bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    apply_widths(ws, {"A": 22, "B": 8, "C": 12, "D": 12, "E": 40, "F": 10})


def build_sources_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("来源与假设")
    ws.append(["类型", "内容", "链接/依据"])
    style_header(ws, 1, fill="E8F5E9")

    rows = [
        ["本地模型尺寸", "你的模型整机约 8.263 x 11.928 x 51.409 cm；单套实体体积约 302.70 cm^3。", "本地几何计算（归档/Y*.stl）"],
        ["打印成本估算", "采用官方国内站 SO-ARM100 无 3D 打印套餐与 3D 打印套餐的价差 450.46 RMB 作为单套第三方打印预算基准。", "https://www.waveshare.net/left_column/Raspberry-Pi-Robotics.htm"],
        ["舵机选择", "SO-ARM100 系列采用 7.4V 19.5kg.cm 总线舵机，12 位磁编码器。", "https://www.waveshare.net/shop/SO-ARM100-CAM-Kit.htm"],
        ["官方 CN BOM", "单臂 682.23 RMB，双臂 1343.16 RMB，用于校验国内自购零件成本。", "https://raw.githubusercontent.com/TheRobotStudio/SO-ARM100/main/SO100.md"],
        ["相机推荐", "罗技 C270 在国内电商长期属于便宜、稳定、即插即用 UVC 方案，近一年常见价格约 89 到 98 RMB。", "https://detail.zol.com.cn/webcams/index256972.shtml ; https://finance.sina.com.cn/tech/roll/2025-05-25/doc-inextxhe4070053.shtml"],
        ["具身完整定义", "这里把“完整具身”定义为 leader/follower 双臂 + 双视角采集 + 现有 PC 可跑数据采集链路。", "基于 LeRobot / SO-100 公开方案的工程定义"],
    ]

    for row in rows:
        ws.append(row)
        if row[2].startswith("http"):
            add_hyperlink(ws[f"C{ws.max_row}"], row[2].split(" ; ")[0], "来源")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    apply_widths(ws, {"A": 16, "B": 60, "C": 40})


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    build_summary_sheet(wb)
    build_model_sheet(wb)
    build_hardware_sheet(wb)
    build_single_arm_sheet(wb)
    build_dual_arm_sheet(wb)
    build_sources_sheet(wb)
    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH.resolve())


if __name__ == "__main__":
    main()
