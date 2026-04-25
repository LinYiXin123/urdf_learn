from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


OUTPUT_PATH = Path("output/spreadsheet/自有模型_最低成本双臂具身_执行包.xlsx")

USED_P2S_LOW = 3000.00
USED_P2S_MID = 3500.00
USED_P2S_HIGH = 4000.00
SERVO_UNIT = 97.72
CONTROL_BOARD_UNIT = 27.00
POWER_UNIT = 22.31
USB_C_UNIT = 23.90
CLAMP_UNIT = 5.20
CAMERA_UNIT = 98.00
SCREWDRIVER_UNIT = 14.90
FASTENER_PACK_UNIT = 50.00
PLA_PLUS_UNIT = 79.00


def style_header(ws, row: int, fill: str = "D9EAF7") -> None:
    thin = Side(style="thin", color="D0D7DE")
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.font = Font(bold=True, color="1F2328")
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)


def style_sheet(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_link(cell, url: str, label: str = "链接") -> None:
    cell.value = label
    cell.hyperlink = url
    cell.style = "Hyperlink"


def build_overview(ws) -> None:
    ws.title = "项目总览"
    ws.append(["项目", "数值", "说明"])
    style_header(ws, 1, "DFF6DD")

    rows = [
        ["项目路线", "二手 P2S + 自有 STL + 双臂具身采集原型", "按已确认路线固化"],
        ["打印机预算区间", "3000–4000 RMB", "二手 P2S"],
        ["双臂硬件与耗材总价", 1679.76, "不含打印机、不含现有 Windows PC"],
        ["项目总价下限", 4679.76, "硬件耗材 1679.76 + P2S 3000"],
        ["项目总价中位", 5179.76, "硬件耗材 1679.76 + P2S 3500"],
        ["项目总价上限", 5679.76, "硬件耗材 1679.76 + P2S 4000"],
        ["第一版目标", "双臂 + 双相机 + LeRobot 采集回放", "不把 AI 自主抓取放进第一版交付"],
        ["里程碑 1", "单臂装配与通信完成", "6 个舵机识别、单关节动作正常"],
        ["里程碑 2", "双臂与双相机就绪", "leader/follower 都可控，双相机稳定采图"],
        ["里程碑 3", "LeRobot 采集链路跑通", "至少完成一次 5–10 分钟采集与回放"],
    ]

    for row in rows:
        ws.append(row)
    for cell in ("A4", "A5", "A6"):
        ws[cell].font = Font(bold=True)

    style_sheet(ws, {"A": 18, "B": 18, "C": 54})


def build_procurement(ws) -> None:
    ws.append(["类别", "项目", "规格", "数量", "预算单价(RMB)", "小计(RMB)", "到货必须检查", "采购入口"])
    style_header(ws, 1, "FFF3CD")

    items = [
        ["打印机", "二手 P2S", "0.4 喷嘴，整机可自检，打印时长约 200h 可接受", 1, USED_P2S_MID, None, "能开机、自检、进退料、热床升温、首层成型", "https://bambulab.com/zh/p2s/specs"],
        ["耗材", "PLA+ 1kg", "1.75mm，优先白/灰等浅色，便于观察缺陷", 1, PLA_PLUS_UNIT, None, "真空包装完整，无明显受潮", ""],
        ["执行器", "STS3215 / ST3215 7.4V C001", "45.2 x 24.7 x 35 mm，12 位磁编码器", 12, SERVO_UNIT, None, "每个都带 3-pin 线 + 舵盘 + 舵盘螺丝", "https://item.taobao.com/item.htm?id=712179366565&skuId=5268252241438"],
        ["控制", "总线舵机控制板", "每台臂 1 块", 2, CONTROL_BOARD_UNIT, None, "接口、跳线、供电口完好", "https://detail.tmall.com/item.htm?id=738817173460&skuId=5096283384143"],
        ["供电", "电源适配器", "每台臂 1 个", 2, POWER_UNIT, None, "输出口与控制板匹配", "https://item.taobao.com/item.htm?id=544824248494&skuId=4974994129990"],
        ["连接", "USB-C 数据线", "数据线而非纯充电线", 2, USB_C_UNIT, None, "能传数据", "https://detail.tmall.com/item.htm?id=44425281296&skuId=5611379016222"],
        ["固定", "桌夹", "共 4 个", 4, CLAMP_UNIT, None, "夹紧无明显偏斜", "https://detail.tmall.com/item.htm?id=801399113134&skuId=5633627126649"],
        ["视觉", "UVC 摄像头", "720p30 起步，优先 1080p30，2 个", 2, CAMERA_UNIT, None, "Windows 下能即插即用", "https://www.jd.com/phb/key_6703325379caeabf8ee.html"],
        ["装配", "备用紧固件包", "M3x6、M2.5x4、M3 螺母", 1, FASTENER_PACK_UNIT, None, "规格齐全", ""],
        ["工具", "十字螺丝刀套装", "至少有 #0 / #1", 1, SCREWDRIVER_UNIT, None, "磁性和批头完好", "https://detail.tmall.com/item.htm?id=675684600845&skuId=4856851392176"],
        ["主机", "现有 Windows PC", "至少 16GB 内存，4 个 USB 口更稳", 1, 0.00, None, "可识别串口与相机", ""],
    ]

    for item in items:
        ws.append(item)
        row = ws.max_row
        ws[f"F{row}"] = f"=D{row}*E{row}"
        if item[7]:
            add_link(ws[f"H{row}"], item[7], "入口")

    ws["A13"] = "双臂硬件与耗材总价（不含 PC）"
    ws["F13"] = "=SUM(F3:F11)"
    ws["A14"] = "项目总价下限（P2S=3000）"
    ws["F14"] = "=F13+3000"
    ws["A15"] = "项目总价上限（P2S=4000）"
    ws["F15"] = "=F13+4000"
    ws["A16"] = "回退条件"
    ws["B16"] = "若 P2S 实际成交价 >4000 或试样件长期不通过，则回退为第三方打印 / A1 mini 路线。"

    for cell in ("A13", "A14", "A15", "A16"):
        ws[cell].font = Font(bold=True)

    style_sheet(ws, {"A": 12, "B": 18, "C": 26, "D": 8, "E": 12, "F": 12, "G": 28, "H": 10})


def build_print_record(ws) -> None:
    ws.append(["参数项", "默认值", "本次实际值", "是否通过", "备注"])
    style_header(ws, 1, "E8F5E9")
    rows = [
        ["打印机", "P2S", "", "", ""],
        ["耗材", "PLA+ 1.75mm", "", "", ""],
        ["喷嘴", "0.4mm", "", "", ""],
        ["层高", "0.2mm", "", "", ""],
        ["填充", "15%", "", "", ""],
        ["打印板", "纹理 PEI", "", "", ""],
        ["首层状态", "均匀连续", "", "", ""],
        ["20mm 测试块 X", "20.0±0.2mm", "", "", ""],
        ["20mm 测试块 Y", "20.0±0.2mm", "", "", ""],
        ["试样件", "先打 Y2.stl", "", "", "优先用来验证舵机仓、孔位和舵盘干涉"],
        ["正式第一套", "试样通过后再打", "", "", ""],
        ["正式第二套", "复用同参数", "", "", ""],
    ]
    for row in rows:
        ws.append(row)
    style_sheet(ws, {"A": 18, "B": 18, "C": 16, "D": 10, "E": 40})


def build_fit_validation(ws) -> None:
    ws.append(["检查项", "通过标准", "结果", "备注"])
    style_header(ws, 1, "FDE2E4")
    rows = [
        ["舵机安装位嵌入", "STS3215 可平稳装入，不需暴力压入", "", ""],
        ["螺丝孔通畅", "M3x6 / M2.5x4 可顺利进入", "", ""],
        ["舵盘安装", "舵盘能安装且不明显偏斜", "", ""],
        ["活动范围", "试样件与舵盘之间无明显干涉", "", ""],
        ["线材出口", "3-pin 线有出口且不夹线", "", ""],
        ["首个完整关节", "装好后可单独转动，无明显卡死", "", ""],
        ["单臂整机", "6 舵机可识别，夹爪能开合", "", ""],
        ["双臂整机", "leader / follower 均可控", "", ""],
        ["双相机", "2 个相机都能采图且编号固定", "", ""],
        ["LeRobot 采集", "可完成一次 5–10 分钟数据采集", "", ""],
    ]
    for row in rows:
        ws.append(row)
    style_sheet(ws, {"A": 20, "B": 34, "C": 10, "D": 32})


def build_servo_id(ws) -> None:
    ws.append(["机械臂", "逻辑关节名", "目标 ID", "实设 ID", "备注"])
    style_header(ws, 1, "D9EAF7")
    follower = [
        ("follower", "shoulder_pan", 1),
        ("follower", "shoulder_lift", 2),
        ("follower", "elbow_flex", 3),
        ("follower", "wrist_flex", 4),
        ("follower", "wrist_roll", 5),
        ("follower", "gripper", 6),
    ]
    leader = [
        ("leader", "shoulder_pan", 1),
        ("leader", "shoulder_lift", 2),
        ("leader", "elbow_flex", 3),
        ("leader", "wrist_flex", 4),
        ("leader", "wrist_roll", 5),
        ("leader", "gripper", 6),
    ]
    for arm, joint, target in follower + leader:
        ws.append([arm, joint, target, "", "每次只接一台臂进行 setup"])
    style_sheet(ws, {"A": 12, "B": 18, "C": 10, "D": 10, "E": 26})


def build_camera_map(ws) -> None:
    ws.append(["相机角色", "Windows 设备索引", "LeRobot 名称", "分辨率", "FPS", "固定位置", "备注"])
    style_header(ws, 1, "FFF3CD")
    ws.append(["front", "", "front", "1280x720", 30, "正前方俯视工作区", "优先固定不动"])
    ws.append(["side", "", "side", "1280x720", 30, "侧前方观察夹爪", "优先固定不动"])
    style_sheet(ws, {"A": 12, "B": 14, "C": 14, "D": 12, "E": 8, "F": 24, "G": 20})


def build_lerobot_flow(ws) -> None:
    ws.append(["步骤", "命令/动作", "说明"])
    style_header(ws, 1, "E8F5E9")
    rows = [
        ["1", "安装 Miniforge / 创建 conda 环境", "Windows 原生 Python 3.12 路线，先不引入 WSL 复杂度"],
        ["2", "conda create -y -n lerobot python=3.12", "创建环境"],
        ["3", "conda activate lerobot", "激活环境"],
        ["4", "conda install ffmpeg -c conda-forge", "安装 ffmpeg"],
        ["5", "pip install \"lerobot[feetech]\"", "安装 LeRobot + Feetech 支持"],
        ["6", "lerobot-find-port", "识别两块控制板的串口"],
        ["7", "lerobot-setup-motors --robot.type=so100_follower --robot.port=COM_FOLLOWER", "按提示依次为 follower 6 个关节设置 ID"],
        ["8", "lerobot-setup-motors --teleop.type=so100_leader --teleop.port=COM_LEADER", "按提示依次为 leader 6 个关节设置 ID"],
        ["9", "lerobot-calibrate --robot.type=so100_follower --robot.port=COM_FOLLOWER --robot.id=follower_arm", "标定 follower"],
        ["10", "lerobot-calibrate --teleop.type=so100_leader --teleop.port=COM_LEADER --teleop.id=leader_arm", "标定 leader"],
        ["11", "huggingface-cli login --token <TOKEN> --add-to-git-credential", "可选，但推荐提前登录"],
        ["12", "lerobot-record --robot.type=so100_follower --robot.port=COM_FOLLOWER --robot.id=follower_arm --robot.cameras=\"{ front: {type: opencv, index_or_path: 0, width: 1280, height: 720, fps: 30}, side: {type: opencv, index_or_path: 1, width: 1280, height: 720, fps: 30}}\" --teleop.type=so100_leader --teleop.port=COM_LEADER --teleop.id=leader_arm --dataset.repo_id=<HF_USER>/so100_custom_record --dataset.num_episodes=5 --dataset.single_task=\"pick and place a cube\" --display_data=true", "第一次录 5 个 episode 足够验链路"],
        ["13", "lerobot-replay --robot.type=so100_follower --robot.port=COM_FOLLOWER --robot.id=follower_arm --dataset.repo_id=<HF_USER>/so100_custom_record --dataset.episode=0", "回放第一个 episode"],
    ]
    for row in rows:
        ws.append(row)
    style_sheet(ws, {"A": 8, "B": 88, "C": 28})


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    build_overview(wb.active)
    build_procurement(wb.create_sheet("最终采购清单"))
    build_print_record(wb.create_sheet("打印参数记录"))
    build_fit_validation(wb.create_sheet("试装验收"))
    build_servo_id(wb.create_sheet("舵机ID表"))
    build_camera_map(wb.create_sheet("双相机映射"))
    build_lerobot_flow(wb.create_sheet("LeRobot采集流程"))

    for ws in wb.worksheets:
        for col in ("E", "F"):
            if ws.title in {"最终采购清单", "项目总览"}:
                for cell in ws[col]:
                    if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                        cell.number_format = "0.00"
        if ws.title == "项目总览":
            for col in ("B",):
                for cell in ws[col]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.00"

    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH.resolve())


if __name__ == "__main__":
    main()
